from __future__ import annotations

from datetime import datetime, date
from pathlib import Path
from collections.abc import Iterable

import pandas as pd
from openpyxl import load_workbook

EXCEL_SUFFIXES = {".xlsx", ".xls"}
DELIMITER = ";"


def is_excel_file(path: Path) -> bool:
    return path.suffix.lower() in EXCEL_SUFFIXES


def read_csv_lines(path: Path) -> list[str]:
    with path.open("r", encoding="utf-8") as handle:
        return [line.rstrip("\n\r") for line in handle]


def stringify_value(value: object) -> str:
    if value is None:
        return ""

    if isinstance(value, str):
        return value

    if isinstance(value, (datetime, date)):
        return value.isoformat()

    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return str(value)

    return str(value)


def _read_xlsx(path: Path) -> list[str]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    lines: list[str] = []
    try:
        for worksheet in workbook.worksheets:
            for row in worksheet.iter_rows(values_only=True):
                stringified = [stringify_value(cell) for cell in row]
                lines.append(DELIMITER.join(stringified))
    finally:
        workbook.close()
    return lines


def _read_xls(path: Path) -> list[str]:
    df: pd.DataFrame = pd.read_excel(path)
    lines: list[str] = []
    for _, row in df.iterrows():
        # Type: ignore pandas cell values as they can be any type
        row_values: list[object] = list(row)
        stringified = [stringify_value(cell) for cell in row_values]
        lines.append(DELIMITER.join(stringified))
    return lines


def read_excel_lines(path: Path) -> list[str]:
    suffix = path.suffix.lower()
    if suffix == ".xls":
        return _read_xls(path)
    if suffix == ".xlsx":
        return _read_xlsx(path)
    raise ValueError(f"Unsupported Excel extension: {suffix}")


def write_csv_lines(path: Path, lines: Iterable[str]) -> None:
    with path.open("w", encoding="utf-8", newline="\n") as handle:
        for line in lines:
            _ = handle.write(line)
            _ = handle.write("\n")
